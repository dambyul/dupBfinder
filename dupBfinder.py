import sys
import mysql.connector
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5 import uic
from multiprocessing import Process, Queue, JoinableQueue
import multiprocessing as mp
import datetime
import time
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from qt_material import apply_stylesheet
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.alert import Alert
import re


__author__ = "LeeDongHwan <gonyang@kakao.com>"

#selenium 옵션 설정
options = webdriver.ChromeOptions() 
options.add_argument("--ignore-certificate-errors") #SSL 인증서 오류 스킵 옵션
options.add_argument('window-size=1920,1080') #창 크기(세로 길이 짧으면 검색 결과 짤림)

def resource_path(relative_path): #exe 파일 만들 때 ui 파일 합치기용
    """ Get absolute path to resource, works for dev and for PyInstaller """
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

def createFolder(directory) : #폴더 생성 함수
    try :
        if not os.path.exists(directory) :
            os.makedirs(directory)
    except OSError :
        print ("폴더 생성 오류 : " + directory)

def makeFolder() :
    # 폴더 생성
    now = datetime.datetime.now()
    current_time = now.strftime("%Y-%m-%d_%H%M%S") #0000-00-00_000000
    createFolder(current_time) #현재 시간으로 폴더 생성
    #조회 결과 폴더 생성
    match = "./"+current_time+"/복본" 
    nomatch = "./"+current_time+"/결과없음"
    createFolder(match)
    createFolder(nomatch)
    #결과 있을 때 사용할 폴더와 없을 때 사용할 폴더 주소를 배열 형태로 return
    return [match,nomatch]

def getData(fn) : #excel 파일 데이터 수집 함수
    #아래 배열과 일치하는 header를 검색하여 데이터 수집
    h_title = ['서명', 'content_title', 'title','책제목','도서명','상품명']
    h_author = ['저자', 'author_name', 'author','저자명','저자및역자','저자 및 역자']
    wb = openpyxl.load_workbook(fn)
    #활성화 되어있는 시트 사용
    sheet = wb.active
    # (1,1) 좌표부터 검색 시작
    header_row,select_col = 1,1
    c_title, c_author = None,None
    # 제목과 저자에 해당되는 열을 찾을 때까지 반복
    while c_title == None or c_author == None :
        #(1,1)부터 검색 시작
        a = sheet.cell(row=header_row, column=select_col).value
        #제목 헤더 검색
        if a in h_title :
            c_title = select_col
        #저역자 헤더 검색
        elif a in h_author :
            c_author = select_col
        #발견못하면 가로 한칸 이동
        select_col += 1
        #100열 넘어가면 행 +1하고 열 1로 초기화
        if select_col >= 100 :
            header_row += 1
            select_col = 1
            #열 10넘어가면 검색 불가능 판정
            if header_row >= 10 :
                break
    list_row = []
    list_title = []
    list_author = []
    start_row = header_row +1
    while 1 :
        # 배열에 데이터의 행 위치와 서명, 저자 데이터를 저장
        if sheet.cell(start_row,c_title).value != None :
            text = str(sheet.cell(start_row,c_title).value)
            text_author = sheet.cell(start_row,c_author).value
            list_row.append(start_row)
            list_title.append(text)
            list_author.append(text_author)
            start_row += 1
        else :
            break
    # 모두 담은 데이터를 return
    data = [list_row,list_title,list_author]
    return data

def producer(q,fn): #excel 파일 읽기
    #excel 파일 불러와서 읽은 후 총 데이터의 개수를 put 해준다
    #에러 발생할 경우 0을 put
    try :
        getResult = getData(fn)
        q.put(len(getResult[0]))
        q.put(getResult)
    except :
       q.put(0)

def producer2(q,object): #복본 조사 및 progress bar 갱신
    ifMultiTask = object[0] #멀티프로세싱 여부
    ifUseHeader = object[1] #headless 여부
    url = object[2] #검색 url
    data = object[3] #검색 데이터 (행 위치, 서명, 저자)
    solution = object[4] #솔루션 데이터
    excelRow = data[0] #행 위치
    excelTitle = data[1] #서명
    excelAuthor = data[2] #저자
    rowNum = len(excelRow) #데이터 개수
    folder = makeFolder() #폴더 생성 및 폴더 좌표값
    if ifUseHeader == False : #headless 체크값 false일 경우
        options.add_argument('headless') #headless 옵션 추가 (웹 페이지 창 안뜸)
    if ifMultiTask == True : #싱글프로세싱
        driver = webdriver.Chrome('chromedriver.exe', options=options) #드라이버 생성 (페이지 열림)
        driver.implicitly_wait(1) #로딩 시간 초과되면 컷
        for i in range(0,rowNum) :
            search(url,solution,folder,driver,excelTitle[i])
            q.put(rowNum)
        driver.quit()
    else : #멀티프로세싱
        print("멀티프로세싱 구현해야함")
    
def search(url,solution,folder,driver,keyword) :
    match = folder[0]
    nomatch = folder[1]
    keyword_xpath = solution[2] #검색 영역
    type_xpath = solution[3] #검색 옵션 (드랍다운)
    type_value = solution[4] #검색 옵션 value
    submit_xpath = solution[5] #검색버튼
    find_result_xpath = solution[6] #검색 결과
    maindata_xpath = solution[7] #첫 컨텐츠 도서 제목
    subdata_xpath = solution[8] #첫 컨텐츠 도서 저자

    driver.get(url) #url 접속
    try :
        Alert(driver).dismiss() #알림창 있을 경우 알림창 종료 (교보 type3 등)
    except :
        pass
    driver.find_element_by_xpath(keyword_xpath).send_keys(keyword) #검색 텍스트 영역, 검색 텍스트
    if type_value: #서명으로 검색 옵션이 활성화 된 경우에만
        select = Select(driver.find_element_by_xpath(type_xpath))
        select.select_by_value(type_value) #value값으로 select 변경
    driver.find_element_by_xpath(submit_xpath).click() #검색 버튼
    try :
        result = True
        print(driver.find_element_by_xpath(maindata_xpath).text.strip()) #첫 컨텐츠 서명 텍스트
        print(driver.find_element_by_xpath(subdata_xpath).text.strip()) #첫 컨텐츠 저자&출판사 텍스트
        #여기 저자 비교 하는거 만들어야함
    except : 
        result = False
        print("검색결과 없음!")
        #결과 없을 경우 서명 잘라서 재검색 하는거 만들어야함

    #검색결과 이미지
    element = driver.find_element_by_xpath(find_result_xpath)
    keyword = re.sub('[\/:*?"<>|]','',keyword)
    if result :
        element.screenshot(match+'/'+keyword+'.png') #검색 결과에 맞는 폴더에 검색어로 저장
        return True
    else :
        element.screenshot(nomatch+'/'+keyword+'.png') #검색 결과에 맞는 폴더에 검색어로 저장
        return False

class Consumer(QThread): #signal 전달 - excel
    poped = pyqtSignal(str)
    excelArr = pyqtSignal(object)
    total_num = pyqtSignal(int)

    def __init__(self, q):
        super().__init__()
        self.q = q

    def run(self):
        while True:
            if not self.q.empty():
                data = q.get()
                if type(data) is int:
                    self.poped.emit("finish")
                    self.total_num.emit(data)
                else :
                    self.excelArr.emit(data)
 
class Consumer2(QThread): #signal 전달 - progress bar
    progressBar = pyqtSignal(float)

    def __init__(self, q2):
        super().__init__()
        self.q2 = q2
        self.cnt = 0

    def run(self):
        while True:
            if not self.q2.empty():
                self.cnt += 1
                data = q2.get()
                self.progressBar.emit((self.cnt/data)*100)            
                    
#gui
form = resource_path('test.ui')
form_class = uic.loadUiType(form)[0] 
class MyWindow(QMainWindow,form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setAcceptDrops(True)

        outdb = mysql.connector.connect(
            host=host,
            user=user,
            password=password,
            port=port
        )
        out_cursor = outdb.cursor(dictionary=True)
        sql = "use hongji"
        out_cursor.execute(sql)

        sql = "select * from dupBFinder order by sort"
        out_cursor.execute(sql)
        
        global fullData
        fullData = []
        for x in out_cursor:
            DBdata = []
            DBdata.append(x['organ_name'])
            DBdata.append(x['url'])
            DBdata.append(x['keyword_xpath'])
            DBdata.append(x['type_xpath'])
            DBdata.append(x['type_value'])
            DBdata.append(x['submit_xpath'])
            DBdata.append(x['find_result_xpath'])
            DBdata.append(x['maindata_xpath'])
            DBdata.append(x['subdata_xpath'])
            fullData.append(DBdata)

        for x in fullData :
            self.comboBox.addItem(x[0]) 

        self.comboBox.activated[str].connect(self.ComboBoxEvent)
        self.pushButton.clicked.connect(self.btn_click)
        #
        # thread for data consumer
        self.consumer = Consumer(q)
        self.consumer2 = Consumer2(q2)
        self.consumer.poped.connect(self.check_ifend)
        self.consumer.total_num.connect(self.event_cnt)
        self.consumer.excelArr.connect(self.getExcelArr)
        self.consumer2.progressBar.connect(self.print_progress)
        self.consumer.start()
        self.consumer2.start()
        

    def ComboBoxEvent(self):
        url = fullData[self.comboBox.currentIndex()][1]
        if url != "" :
            self.lineEdit.setText(url)

    def getExcelArr(self,data):
        global getArr
        getArr = data

    def event_cnt(self,data) :
        if data == 0 :
            QMessageBox.about(self,'엑셀 불러오기 실패',"검색된 도서가 없습니다. 양식을 확인해주세요.")
            self.label.setText("엑셀 파일 (.xlsx)을 프로그램으로 드래그 해주세요")
        else :
            global getArrCnt
            getArrCnt = data
            QMessageBox.about(self,'엑셀 불러오기 완료',str(data) + "종의 도서를 불러왔습니다.")
            self.label.setText(self.label.text() + " - " + str(data) + "종")


    @pyqtSlot(float)
    def print_progress(self, data):
        self.progressBar.setValue(data)

    @pyqtSlot(str)
    def check_ifend(self, data):
        if data == "start" :
            self.widget.setEnabled(False)
        elif data == "finish" :
            self.widget.setEnabled(True)

    def btn_click(self): 
        try :
            arr = getArr
            pros = self.radioButton.isChecked()
            test = self.checkBox.isChecked()
            url = self.lineEdit.text()
            if url == "" :
                QMessageBox.critical(self,'필수 데이터 부족!','URL이 입력되지 않았습니다.')
            else : 
                p2 = Process(name="producer2", target=producer2, args=(q2, [pros,test,url,arr,fullData[self.comboBox.currentIndex()]]), daemon=True)
                p2.start()
        except :
            QMessageBox.critical(self,'필수 데이터 부족!','엑셀 파일을 업로드해주세요.')

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        global filename
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for filename in files:
            print("drop : " + filename)
        fn = filename.split("/")
        fn = fn[len(fn)-1]
        self.label.setText(fn)
        p = Process(name="producer", target=producer, args=(q,filename ), daemon=True)
        self.check_ifend("start")
        p.start()
        

if __name__ == "__main__":
    q = Queue()
    q2 = Queue()
    # producer process
    pool = mp.Pool(processes=4)
    # Main process
    app = QApplication(sys.argv)
    mywindow = MyWindow()
    apply_stylesheet(app, theme='light_blue.xml', invert_secondary=True)
    mywindow.show()
    app.exec_()